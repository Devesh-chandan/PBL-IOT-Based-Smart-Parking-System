from fastapi import FastAPI, File, UploadFile
from ultralytics import YOLO
import easyocr
import openpyxl
import cv2
import numpy as np
from datetime import datetime
from difflib import SequenceMatcher
import os
import re

app    = FastAPI()
model  = YOLO("../model/weights/best.pt")
reader = easyocr.Reader(['en'])
EXCEL  = "parking_log.xlsx"

def init_excel():
    if not os.path.exists(EXCEL):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Parking Log"
        ws.append(["ID", "Plate Number", "Entry Time", "Exit Time", "Status"])
        wb.save(EXCEL)

def clean_plate(text):
    text = re.sub(r'[^A-Z0-9]', '', text.upper())
    if text.startswith('HH'):
        text = 'M' + text[2:]
    if text.startswith('H') and len(text) > 1 and not text.startswith('HH'):
        text = 'M' + text[1:]
    text = text.replace('I', '1').replace('O', '0').replace('S', '5').replace('B', '8')
    return text

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def find_plate_row(ws, plate):
    best_match_row = None
    best_ratio     = 0.0

    for row in ws.iter_rows(min_row=2):
        if row[4].value == "INSIDE" and row[1].value:
            ratio = similar(plate, row[1].value)
            if ratio > best_ratio:
                best_ratio     = ratio
                best_match_row = row[0].row

    if best_ratio >= 0.7:
        return best_match_row
    return None

@app.post("/upload")
async def upload(lane: str, image: UploadFile = File(...)):
    init_excel()

    contents = await image.read()
    nparr    = np.frombuffer(contents, np.uint8)
    img      = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    temp_path = "temp_plate.jpg"
    cv2.imwrite(temp_path, img)

    results = model(temp_path, conf=0.5)
    if len(results[0].boxes) == 0:
        return {"status": "NO_PLATE", "message": "No plate detected"}

    box          = results[0].boxes[results[0].boxes.conf.argmax()]
    x1,y1,x2,y2 = map(int, box.xyxy[0])
    crop         = img[y1:y2, x1:x2]

    ocr_result = reader.readtext(crop)
    if not ocr_result:
        return {"status": "OCR_FAILED", "message": "Could not read plate"}

    plate_text = clean_plate(ocr_result[0][1])
    now        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active

    if lane == "entry":
        new_id = ws.max_row
        ws.append([new_id, plate_text, now, "", "INSIDE"])
        wb.save(EXCEL)
        return {"status": "OPEN_GATE", "plate": plate_text, "time": now}

    elif lane == "exit":
        row_num = find_plate_row(ws, plate_text)
        if row_num:
            ws.cell(row=row_num, column=4).value = now
            ws.cell(row=row_num, column=5).value = "EXITED"
            wb.save(EXCEL)
            return {"status": "OPEN_GATE", "plate": plate_text, "time": now}
        else:
            return {"status": "NOT_FOUND", "message": "No entry record found"}

    return {"status": "ERROR"}

@app.get("/records")
def get_records():
    init_excel()
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            records.append({
                "id"         : row[0],
                "plate"      : row[1],
                "entry_time" : row[2],
                "exit_time"  : row[3],
                "status"     : row[4]
            })
    return {"records": records}